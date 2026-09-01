import io
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def mask_text(val, keep_front=1, keep_back=1):
    if not val or val == "-":
        return "-"
    s = str(val).strip()
    if len(s) <= keep_front + keep_back:
        return "*" * len(s)
    return s[:keep_front] + "*" * (len(s) - keep_front - keep_back) + s[-keep_back:]

def mask_nik(nik):
    if not nik or nik == "-":
        return "-"
    s = str(nik).strip()
    if len(s) >= 8:
        return s[:4] + "********" + s[-4:]
    return "******"

def mask_rm(rm):
    if not rm or rm == "-":
        return "-"
    s = str(rm).strip()
    if len(s) >= 4:
        return s[:2] + "****" + s[-2:]
    return "****"

def generate_research_excel(records: list, filters: dict, anonymize: bool = True) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Penelitian SIHA"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy/Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_font = Font(name="Calibri", size=10, italic=True, color="475569")
    bold_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    # 1. Header Information / Metadata Block
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "LAPORAN EKSTRAKSI DATA PENELITIAN HIV-AIDS (SIHA RS)"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")

    ws["A2"] = "Waktu Ekstraksi:"
    ws["B2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"].font = bold_font
    ws["B2"].font = regular_font

    ws["A3"] = "Status Privasi:"
    ws["B3"] = "TER-DEIDENTIFIKASI / ANONIM (Sesuai UU PDP)" if anonymize else "DATA ASLI (Akses Terbatas)"
    ws["A3"].font = bold_font
    ws["B3"].font = Font(name="Calibri", size=10, bold=True, color="10B981" if anonymize else "EF4444")

    ws["A4"] = "Total Data Pasien:"
    ws["B4"] = len(records)
    ws["A4"].font = bold_font
    ws["B4"].font = bold_font

    start_row = 6

    # 2. Table Column Definitions
    columns = [
        ("No", 6),
        ("ID Pasien", 14),
        ("No Rekam Medik", 16),
        ("Nama Pasien", 20),
        ("NIK / Identitas", 18),
        ("Usia", 8),
        ("Kategori Usia", 14),
        ("Klasifikasi Usia WHO", 20),
        ("Jenis Kelamin", 14),
        ("Status Asal Pasien (Asli/Rujukan)", 26),
        ("Status Retensi PDP", 22),
        ("Status Hamil", 14),
        ("Kelompok Populasi", 22),
        ("Domisili Kab/Kota", 20),
        ("Tgl Register", 14),
        ("Tgl Kunjungan Terakhir", 18),
        ("Alasan Kunjungan", 22),
        ("Kepatuhan (Keterlambatan)", 22),
        ("Stadium WHO", 14),
        ("Nama Rejimen ARV", 28),
        ("Kategori Rejimen", 20),
        ("Jml Hari ARV", 14),
        ("Status MMD (>=60 hr)", 18),
        ("Koinfeksi TB", 18),
        ("Berat Badan (kg)", 14),
        ("Tinggi Badan (cm)", 14),
        ("IMT", 10),
        ("Kategori IMT", 18),
        ("Akhir Follow Up", 16),
        ("Durasi LTFU", 18),
        ("Tgl Tes VL", 14),
        ("Hasil VL", 18),
        ("Kategori Viral Load", 24),
        ("Status Supresi", 16),
        ("Tgl Tes CD4", 14),
        ("Nilai CD4 (sel/uL)", 16),
        ("Kategori CD4", 24)
    ]

    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, len(col_name) + 4)

    ws.row_dimensions[start_row].height = 28

    # 3. Data Rows
    current_row = start_row + 1
    for idx, r in enumerate(records, 1):
        p_id = r["pasien_id"]
        rm = mask_rm(r["no_rekam_medik"]) if anonymize else r["no_rekam_medik"]
        nama = mask_text(r["nama_pasien"], 1, 1) if anonymize else r["nama_pasien"]
        nik = mask_nik(r["nik"]) if anonymize else r["nik"]
        
        row_values = [
            idx,
            p_id,
            rm,
            nama,
            nik,
            r["umur"],
            r["kategori_umur"],
            r.get("kelompok_umur_klinis", "-"),
            r["jenis_kelamin"],
            "Transfer-In Rujukan (Faskes Luar)" if r.get("is_rujuk_masuk") else "Inisiasi Internal (RS Kariadi)",
            r.get("status_pdp", "-"),
            r.get("status_hamil", "-"),
            r["kelompok_populasi"],
            r["domisili_kabupaten"],
            r["tanggal_register"],
            r["tanggal_kunjungan"],
            r["alasan_kunjungan"],
            r.get("kepatuhan", "-"),
            r["stadium_klinis"],
            r["nama_rejimen"],
            r["kategori_rejimen"],
            r["jumlah_hari_arv"],
            "Ya (MMD)" if r.get("is_mmd") else "Reguler (<60 hr)",
            r.get("koinfeksi_tb", "-"),
            r["berat_badan"],
            r["tinggi_badan"],
            r["imt"],
            r["kategori_imt"],
            r["akhir_follow_up"],
            r.get("ltfu_kategori", "-"),
            r["tanggal_vl"],
            r["hasil_vl_raw"],
            r["kategori_vl"],
            "Tersupresi" if r["is_suppressed"] else ("Gagal Virologis" if "Gagal" in r["kategori_vl"] else "-"),
            r.get("tanggal_cd4"),
            r.get("nilai_cd4"),
            r.get("kategori_cd4")
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val if val is not None else "-")
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 6, 17, 18, 19, 20, 27, 28]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [11, 12, 22, 23]:
                cell.alignment = Alignment(horizontal="center")

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_research_csv(records: list, anonymize: bool = True) -> io.BytesIO:
    rows = []
    for idx, r in enumerate(records, 1):
        rows.append({
            "no": idx,
            "pasien_id": r["pasien_id"],
            "no_rekam_medik": mask_rm(r["no_rekam_medik"]) if anonymize else r["no_rekam_medik"],
            "nama_pasien": mask_text(r["nama_pasien"], 1, 1) if anonymize else r["nama_pasien"],
            "nik": mask_nik(r["nik"]) if anonymize else r["nik"],
            "umur": r["umur"],
            "kategori_umur": r["kategori_umur"],
            "kelompok_umur_klinis": r.get("kelompok_umur_klinis", "-"),
            "jenis_kelamin": r["jenis_kelamin"],
            "status_asal_pasien": "Transfer-In Rujukan (Faskes Luar)" if r.get("is_rujuk_masuk") else "Inisiasi Internal (RS Kariadi)",
            "status_pdp": r.get("status_pdp", "-"),
            "status_hamil": r.get("status_hamil", "-"),
            "is_mmd": "Ya (MMD)" if r.get("is_mmd") else "Reguler",
            "kelompok_populasi": r["kelompok_populasi"],
            "domisili_kabupaten": r["domisili_kabupaten"],
            "tanggal_register": r["tanggal_register"],
            "tanggal_kunjungan": r["tanggal_kunjungan"],
            "alasan_kunjungan": r["alasan_kunjungan"],
            "kepatuhan": r.get("kepatuhan", "-"),
            "stadium_klinis": r["stadium_klinis"],
            "nama_rejimen": r["nama_rejimen"],
            "kategori_rejimen": r["kategori_rejimen"],
            "jumlah_hari_arv": r["jumlah_hari_arv"],
            "koinfeksi_tb": r.get("koinfeksi_tb", "-"),
            "berat_badan": r["berat_badan"],
            "tinggi_badan": r["tinggi_badan"],
            "imt": r["imt"],
            "kategori_imt": r["kategori_imt"],
            "akhir_follow_up": r["akhir_follow_up"],
            "durasi_ltfu": r.get("ltfu_kategori", "-"),
            "tanggal_vl": r["tanggal_vl"],
            "hasil_vl": r["hasil_vl_raw"],
            "kategori_vl": r["kategori_vl"],
            "is_suppressed": r["is_suppressed"],
            "tanggal_cd4": r.get("tanggal_cd4"),
            "nilai_cd4": r.get("nilai_cd4"),
            "kategori_cd4": r.get("kategori_cd4")
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    return buf


def generate_pediatric_excel(records: list, period_label: str = "Semua Waktu", filter_type: str = "Semua Kategori", anonymize: bool = False) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kohor Anak & Bayi PPIA"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy/Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    regular_font = Font(name="Calibri", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    # 1. Header Metadata Block
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "REGISTER RINCIAN KOHOR ANAK, BAYI TERPAJAN & PPIA (<18 TAHUN)"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")

    ws["A2"] = "Fasilitas Pelayanan:"
    ws["B2"] = "RSUP Dr. Kariadi Semarang"
    ws["A2"].font = bold_font
    ws["B2"].font = regular_font

    ws["A3"] = "Waktu Unduh:"
    ws["B3"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"].font = bold_font
    ws["B3"].font = regular_font

    ws["A4"] = "Kategori Filter:"
    ws["B4"] = filter_type
    ws["A4"].font = bold_font
    ws["B4"].font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")

    ws["D2"] = "Periode Kunjungan:"
    ws["E2"] = period_label
    ws["D2"].font = bold_font
    ws["E2"].font = regular_font

    ws["D3"] = "Total Data Anak:"
    ws["E3"] = len(records)
    ws["D3"].font = bold_font
    ws["E3"].font = Font(name="Calibri", size=10, bold=True, color="059669")

    ws["D4"] = "Status Data:"
    ws["E4"] = "TER-DEIDENTIFIKASI (Masking)" if anonymize else "DATA ASLI LENGKAP"
    ws["D4"].font = bold_font
    ws["E4"].font = Font(name="Calibri", size=10, bold=True, color="10B981" if anonymize else "2563EB")

    start_row = 6

    # 2. Columns
    columns = [
        ("No", 6),
        ("Pasien ID", 14),
        ("No. Rekam Medik", 16),
        ("Nama Pasien", 24),
        ("NIK / Identitas", 18),
        ("Usia (Tahun)", 12),
        ("Kategori Usia", 14),
        ("Jenis Kelamin", 14),
        ("Tgl Kunjungan Terakhir", 18),
        ("Kategori Status Anak", 32),
        ("Rejimen / Profilaksis", 28),
        ("Alasan Kunjungan", 24),
        ("Tgl Mulai Terapi", 16),
        ("Durasi Terapi (Th)", 16),
        ("Status Retensi PDP", 20),
        ("Hasil Lab VL Terakhir", 20),
        ("Status Supresi VL", 18),
        ("Nilai CD4 Terakhir", 16),
        ("Kategori CD4", 22),
        ("Berat Badan (kg)", 14),
        ("Tinggi Badan (cm)", 14),
        ("IMT", 10),
        ("Kategori IMT", 18),
        ("Domisili Kab/Kota", 20),
        ("Faskes Rujukan Asal", 24)
    ]

    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, len(col_name) + 4)

    ws.row_dimensions[start_row].height = 28

    # 3. Rows
    current_row = start_row + 1
    for idx, r in enumerate(records, 1):
        p_id = r.get("pasien_id", "-")
        rm = mask_rm(r.get("no_rekam_medik")) if anonymize else (r.get("no_rekam_medik") or "-")
        nama = mask_text(r.get("nama_pasien"), 1, 1) if anonymize else (r.get("nama_pasien") or "-")
        nik = mask_nik(r.get("nik")) if anonymize else (r.get("nik") or "-")
        
        status_anak = r.get("status_anak_detail") or ("Anak Terkonfirmasi ODHIV" if r.get("status_odhiv") == "ODHIV" else "Bayi Terpajan Profilaksis")
        
        row_values = [
            idx,
            p_id,
            rm,
            nama,
            nik,
            r.get("umur", "-"),
            r.get("kategori_umur", "Anak"),
            r.get("jenis_kelamin", "-"),
            r.get("tanggal_kunjungan", "-"),
            status_anak,
            r.get("nama_rejimen", "-"),
            r.get("alasan_kunjungan", "-"),
            r.get("tanggal_mulai_art", "-"),
            r.get("durasi_art_tahun", "-"),
            r.get("status_pdp", "-"),
            r.get("hasil_vl_raw", "-"),
            "Tersupresi" if r.get("is_suppressed") else ("Gagal Virologis" if "Gagal" in str(r.get("kategori_vl", "")) else "-"),
            r.get("nilai_cd4", "-"),
            r.get("kategori_cd4", "-"),
            r.get("berat_badan", "-"),
            r.get("tinggi_badan", "-"),
            r.get("imt", "-"),
            r.get("kategori_imt", "-"),
            r.get("domisili_kabupaten", "-"),
            r.get("asal_rujukan", "-")
        ]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val if val is not None else "-")
            cell.font = regular_font
            cell.border = thin_border
            if col_idx in [1, 6, 8, 9, 13, 14, 17, 18, 20, 21, 22]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [2, 3, 5]:
                cell.alignment = Alignment(horizontal="center")

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
